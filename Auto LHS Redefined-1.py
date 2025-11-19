# -*- coding: utf-8 -*-
"""
Created on Fri Mar  7 20:38:49 2025

@author: 20353120
"""

import string
from openpyxl import load_workbook
from tkinter import Tk, Label, Entry, Button, messagebox
# Function to convert Excel column letter to zero-based index
def excel_col_to_index(col):
   col = col.upper()
   index = 0
   for i, char in enumerate(reversed(col)):
       index += (string.ascii_uppercase.index(char) + 1) * (26 ** i)
   return index
def process_excel():
   try:
       # Get inputs from the user
       loop_number = loop_number_entry.get()
       standard_format_path = 'C:/Users/20353120/Downloads/Master Line History Sheet (LHS).xlsx'
       master_sheet_path = "D:/OneDrive - Larsen & Toubro/Desktop/Erection Piping LHS.xlsx"
       output_path = f'C:/Users/20353120/Downloads/Master Line History Sheet (LHS) _ {loop_number}.xlsx'
       # Validate inputs
       if not loop_number:
           messagebox.showerror("Error", "Please enter a loop number.")
           return
       # Load the standard format and master sheet
       standard_wb = load_workbook(standard_format_path)
       master_ws = load_workbook(master_sheet_path, data_only=True)['L.H.S.']  # Adjust sheet name
       # Get the active sheet for standard format
       standard_sheet = standard_wb.active
       # Step 1: Populate the loop number in the standard format
       standard_sheet.cell(row=4, column=14).value = loop_number  # Example: Cell N4 (row 4, col 14)
       # Step 2: Filter the master sheet for matching loop number
       matching_rows = []
       for row in master_ws.iter_rows(min_row=2, values_only=True):  # Skip header
           if row[3] == loop_number:  # Assuming "Loop No" is in column A
               matching_rows.append(row)
       if not matching_rows:
           messagebox.showinfo("Info", f"No matching loop number found for {loop_number}.")
           return
       # Step 3: Define column mapping (master sheet to standard format)
       column_mapping = {

           "ISO NO." : {"column": 2, "start_row": 8, "master_col_index": 2},

           "Joint No." : {"column": 3, "start_row": 8, "master_col_index": 16},

           "Spec./Class": {"column": 4, "start_row": 8, "master_col_index": 9},

           "Type of Weld": {"column": 5, "start_row": 8, "master_col_index": 20},

           "Dia": {"column": 6, "start_row": 8, "master_col_index": 18},

           "Thk.": {"column": 7, "start_row": 8, "master_col_index": 19},
           
           "Fitup Clearance Report No.": {"column" : 25, "start_row": 8, "master_col_index": 21},
           
           "Welding Visual Clearance Report": {"column": 27, "start_row": 8, "master_col_index": 23},
           
           "Visual Date": {"column" : 28, "start_row": 8, "master_col_index": 24},
           
           "Welder Stamp": {"column": 10, "start_row": 8, "master_col_index": 25},
           
           "WPS  No & Rev.": {"column": 11, "start_row": 8, "master_col_index": 26},
           
           "DPT Report No.": {"column": 29, "start_row": 8, "master_col_index": 35},
           
           "DPT Date": {"column": 30, "start_row": 8, "master_col_index": 36},
           
           "DPT Group No.": {"column": 13, "start_row": 8, "master_col_index": 33},
           
           "Report No. 1": {"column": 31, "start_row": 8, "master_col_index": 44},
           
           "RT Report Date R0": {"column": 32, "start_row": 8, "master_col_index": 45},
           
           "RT Group No.": {"column": 15, "start_row": 8, "master_col_index": 40},
           
           "PMI Report No.": {"column": 33, "start_row": 8, "master_col_index": 27},
           
           "PMI Date": {"column": 34, "start_row": 8, "master_col_index": 28},
           
           "Ferrite Report No.": {"column": 35, "start_row": 8, "master_col_index": 30},
           
           "Ferrite Date": {"column": 36, "start_row": 8, "master_col_index": 31}
           
       }

       for master_col_index, target_info in column_mapping.items():
           target_col = target_info["column"]
           start_row = target_info["start_row"]
           master_col_index = target_info["master_col_index"] - 1
           for i, row in enumerate(matching_rows):
               value = row[master_col_index]  # Get corresponding master column
               standard_sheet.cell(row=start_row + i, column=target_col).value = value
       # Step 5: Save the updated Excel file
       standard_wb.save(output_path)
       messagebox.showinfo("Success", f"Updated file saved as {output_path}")
   except Exception as e:
       messagebox.showerror("Error", f"An error occurred: {e}")
# Tkinter GUI
root = Tk()
root.title("Loop File LHS Generation Program (LFLGP)")
# Labels and Entry widgets
Label(root, text="Loop Number").grid(row=2, column=0, padx=10, pady=5)
loop_number_entry = Entry(root, width=30)
loop_number_entry.grid(row=2, column=1, padx=10, pady=5)
# Process Button
process_button = Button(root, text="Process Excel", command=process_excel, bg="blue", fg="white")
process_button.grid(row=3, column=0, columnspan=2, pady=20)
# Run the Tkinter event loop
root.mainloop()